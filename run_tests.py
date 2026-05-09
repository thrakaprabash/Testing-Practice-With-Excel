"""
run_tests.py – Read test cases from an XLSX file, translate each input with
Google Translate (Playwright), then write the actual output and PASS/FAIL
status back into the same XLSX file.

Usage:
    python run_tests.py [--xlsx PATH] [--headless] [--slowmo MS]
                        [--timeout MS] [--delay MS]

Columns expected in the XLSX (row 1 = header):
    input | expected_output | actual_output | status
"""

import argparse
import re
import sys
import time
from pathlib import Path

# Force UTF-8 output so Sinhala characters print safely on Windows
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# ── Google Translate constants ──────────────────────────────────────────────
TRANSLATE_URL = (
    "https://translate.google.com/?sl=si&tl=en&hl=en&op=translate"
)
SOURCE_INPUT_SELECTOR = "textarea[aria-label='Source text']"
TARGET_OUTPUT_SELECTORS = (
    "div[data-testid='target-text'], span[jsname='W297wb']"
)

# ── Column indices (1-based, matching header row) ───────────────────────────
COL_INPUT = 1
COL_EXPECTED = 2
COL_ACTUAL = 3
COL_STATUS = 4


# ── Helpers ─────────────────────────────────────────────────────────────────

def clear_and_wait_for_empty(page, input_box, target_locator, timeout_ms: int) -> None:
    """
    Clear the source textarea and wait until the translation panel is blank.
    This prevents the previous translation from being read as the new result.
    """
    input_box.fill("")
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        text = ""
        if target_locator.count() > 0:
            text = target_locator.first.inner_text().strip()
        if not text:
            return
        page.wait_for_timeout(100)


def wait_for_translation(
    page, target_locator, timeout_ms: int, min_wait_ms: int
) -> str:
    """Poll the translation output until it stabilises, then return it."""
    deadline = time.monotonic() + timeout_ms / 1000
    ready_time = time.monotonic() + min_wait_ms / 1000
    last_text = ""
    stable_reads = 0
    stable_reads_required = 3

    while time.monotonic() < deadline:
        text = ""
        if target_locator.count() > 0:
            text = target_locator.first.inner_text().strip()

        if text:
            if time.monotonic() < ready_time:
                last_text = text
                stable_reads = 0
            elif text == last_text:
                stable_reads += 1
                if stable_reads >= stable_reads_required:
                    return text
            else:
                last_text = text
                stable_reads = 0

        page.wait_for_timeout(200)

    return last_text


def expected_options(expected_output: str) -> list[str]:
    """
    Split an expected-output cell into individual acceptable answers.
    Parenthetical notes are stripped, then we split on  /  |  or.
    """
    cleaned = re.sub(r"\s*\([^)]*\)", "", expected_output).strip()
    if not cleaned:
        return []
    parts = re.split(r"\s*/\s*|\s*\|\s*|\s+or\s+", cleaned, flags=re.IGNORECASE)
    return [p.strip() for p in parts if p.strip()]


# Common English contractions → expanded forms for normalisation
_CONTRACTIONS = [
    (r"i'm\b",      "i am"),
    (r"i'll\b",     "i will"),
    (r"i've\b",     "i have"),
    (r"i'd\b",      "i would"),
    (r"you're\b",   "you are"),
    (r"you'll\b",   "you will"),
    (r"you've\b",   "you have"),
    (r"he's\b",     "he is"),
    (r"she's\b",    "she is"),
    (r"it's\b",     "it is"),
    (r"we're\b",    "we are"),
    (r"they're\b",  "they are"),
    (r"can't\b",    "cannot"),
    (r"won't\b",    "will not"),
    (r"don't\b",    "do not"),
    (r"didn't\b",   "did not"),
    (r"doesn't\b",  "does not"),
    (r"isn't\b",    "is not"),
    (r"aren't\b",   "are not"),
    (r"wasn't\b",   "was not"),
    (r"weren't\b",  "were not"),
    (r"hasn't\b",   "has not"),
    (r"haven't\b",  "have not"),
    (r"hadn't\b",   "had not"),
    (r"couldn't\b", "could not"),
    (r"wouldn't\b", "would not"),
    (r"shouldn't\b","should not"),
    (r"mustn't\b",  "must not"),
    (r"i'm not\b",  "i am not"),
]


def _normalise(text: str) -> str:
    """Lowercase, expand contractions, strip punctuation for fuzzy comparison."""
    t = text.strip().lower()
    for pattern, replacement in _CONTRACTIONS:
        t = re.sub(pattern, replacement, t)
    t = re.sub(r"[^\w\s]", "", t)   # strip punctuation
    t = re.sub(r"\s+", " ", t).strip()
    return t


def is_match(actual_output: str, expected_output: str) -> bool:
    """Return True when actual_output satisfies any of the expected options."""
    if not actual_output or not expected_output:
        return False
    actual_norm = _normalise(actual_output)
    for option in expected_options(expected_output):
        option_norm = _normalise(option)
        if actual_norm == option_norm:
            return True
        # actual contains the expected phrase (e.g. longer translation)
        if option_norm and option_norm in actual_norm:
            return True
        # expected contains the actual phrase (e.g. actual is a substring)
        if actual_norm and actual_norm in option_norm:
            return True
    return False


# ── Core runner ─────────────────────────────────────────────────────────────

def run(
    xlsx_path: Path,
    headless: bool,
    slowmo: int,
    timeout_ms: int,
    delay_ms: int,
) -> None:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Validate header row
    headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    print(f"[INFO] Header row: {headers}")

    # Collect data rows (skip header)
    data_rows = []
    for row_idx in range(2, ws.max_row + 1):
        input_text = ws.cell(row=row_idx, column=COL_INPUT).value
        expected = ws.cell(row=row_idx, column=COL_EXPECTED).value
        if input_text is None:
            continue  # skip blank rows
        data_rows.append((row_idx, str(input_text).strip(), str(expected or "").strip()))

    print(f"[INFO] Found {len(data_rows)} test case(s) to run.\n")

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=headless, slow_mo=slowmo)
        context = browser.new_context()
        page = context.new_page()
        page.set_default_timeout(timeout_ms)

        page.goto(TRANSLATE_URL, wait_until="domcontentloaded")
        input_box = page.locator(SOURCE_INPUT_SELECTOR)
        input_box.wait_for(state="visible")
        target_locator = page.locator(TARGET_OUTPUT_SELECTORS)

        # Warmup: type and clear a dummy char so the page is fully ready
        # before we start the real test cases (avoids empty first result).
        input_box.fill("a")
        page.wait_for_timeout(1000)
        clear_and_wait_for_empty(page, input_box, target_locator, timeout_ms)

        for row_idx, input_text, expected_output in data_rows:
            # Clear old translation first to avoid stale-text reads
            clear_and_wait_for_empty(page, input_box, target_locator, timeout_ms)
            input_box.fill(input_text)
            actual_output = wait_for_translation(
                page, target_locator, timeout_ms, delay_ms
            )
            status = "PASS" if is_match(actual_output, expected_output) else "FAIL"

            # Write results back into the same row
            ws.cell(row=row_idx, column=COL_ACTUAL).value = actual_output
            ws.cell(row=row_idx, column=COL_STATUS).value = status

            print(
                f"  Row {row_idx:>3} | {input_text[:30]:<30} | "
                f"Actual: {actual_output[:30]:<30} | {status}"
            )

        browser.close()

    # Save to a temp file first, then replace the original.
    # This avoids PermissionError if the xlsx is currently open in Excel.
    tmp_path = xlsx_path.with_suffix(".tmp.xlsx")
    wb.save(tmp_path)
    try:
        tmp_path.replace(xlsx_path)
    except PermissionError:
        print(
            f"\n[WARN] Could not overwrite {xlsx_path} (is it open in Excel?).\n"
            f"       Results saved to: {tmp_path}\n"
            f"       Close Excel and rename that file manually."
        )
        return
    print(f"\n[DONE] Results saved to: {xlsx_path}")


# ── CLI ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Translate Sinhala→English with Google Translate and update "
            "PASS/FAIL status directly in the XLSX file."
        )
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path("translate_results.xlsx"),
        help="Path to the XLSX file (read and updated in-place).",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run without a visible browser window.",
    )
    parser.add_argument(
        "--slowmo",
        type=int,
        default=0,
        help="Slow down Playwright actions (ms).",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=15000,
        help="Timeout for UI operations (ms).",
    )
    parser.add_argument(
        "--delay",
        type=int,
        default=2000,
        help="Minimum wait after typing before reading translation (ms).",
    )

    args = parser.parse_args()
    run(args.xlsx, args.headless, args.slowmo, args.timeout, args.delay)


if __name__ == "__main__":
    main()
